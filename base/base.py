# -*- coding: utf-8 -*-# 
# ------------------------------------------------------
# @Time :      2021/1/14 10:10 
# @Author :    Neil.Jiang
# @project:    YiQiaUiTest
# @File :      base.py
# @Email :     jiang.nell@qq.com
# @Description: 
# @命名规范：文件名全小写+下划线，类名大写字母开头，方法和变量小写+下划线连接，常量大写

# -------------------------------------------------------

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import logging
import os
import time
from PIL import Image
import math
import operator
from functools import reduce
import urllib.parse
import openpyxl
from selenium.webdriver.remote.webelement import WebElement
import json

class Base():
    def __init__(self, url, browser="chrome",):
        """
        :param url:登录的url
        :param browser: 选择的浏览器
        """
        if browser in ("chrome", "Chrome"):
            self.driver = webdriver.Chrome()
            self.driver.get(url=url)
            self.driver.maximize_window()
        elif browser in ("firefox", "Firefox"):
            self.driver = webdriver.Firefox()
            self.driver.get(url=url)
            self.driver.maximize_window()
        elif browser in ("ie", "Ie"):
            self.driver = webdriver.Ie()
            self.driver.get(url=url)
        elif browser in ("safari", "Safari"):
            self.driver = webdriver.Safari()
            self.driver.get(url=url)
        elif browser in ("opera", "Opera"):
            self.driver = webdriver.Opera()
            self.driver.get(url=url)
        else:
            raise NameError("您输入的浏览器有误请输入五大浏览器\n列如：chrome,firefox,ie,safari,opera")

    def find_ele(self, ele_type, ele, timeout=5,):
        """
        以显示等待的方式定位元素
        :param ele_type: 元素类型
        :param ele: 元素的具体文本
        :param timeout: 等待时间
        :return:
        """
        if ele_type in ("id", "Id"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.element_to_be_clickable((By.ID, ele)))

        elif ele_type in ("xpath", "Xpath"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.element_to_be_clickable((By.XPATH, ele)))
        elif ele_type in ("name", "Name"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.element_to_be_clickable((By.NAME, ele)))
        elif ele_type in ("class", "Class"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.element_to_be_clickable((By.CLASS_NAME, ele)))
        elif ele_type in ("css", "Css"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.element_to_be_clickable((By.CSS_SELECTOR, ele)))

        elif ele_type in ("text", "Text"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.element_to_be_clickable((By.LINK_TEXT, ele)))

    def find_eles(self, ele_type, ele, timeout=5,):
        """
        以显示等待的方式定位复数元素
        :param ele_type: 元素的类型
        :param ele: 元素的具体文本
        :param timeout: 等待时间
        :return:
        """
        if ele_type in ("id", "Id"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.visibility_of_all_elements_located((By.ID, ele)))
        elif ele_type in ("xpath", "Xpath"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.visibility_of_all_elements_located((By.XPATH, ele)))
        elif ele_type in ("name", "Name"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.visibility_of_all_elements_located((By.NAME, ele)))
        elif ele_type in ("class", "Class"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.visibility_of_all_elements_located((By.CLASS_NAME, ele)))
        elif ele_type in ("css", "Css"):
            return WebDriverWait(self.driver,
                                 timeout=timeout).until(ec.visibility_of_all_elements_located((By.CSS_SELECTOR, ele)))

    def find_frame(self,  frame_content, timeout=5, back_defult=True):
        """
        切换iframe框架
        :param timeout:等待时间
        :param frame_content: frame_content可以为frame的 id,name,index,或WebElement元素
        :param back_defult: 切换到默认iframe
        :return:
        """
        if back_defult:
            self.driver.switch_to.default_content()
            if frame_content == '':
                #  返回到默认content
                pass
            else:
                #  返回到默认content的下一级iframe里面
                WebDriverWait(self.driver, timeout=timeout).until(ec.frame_to_be_available_and_switch_to_it(frame_content))
        else:
            # 前往到当前iframe的下一级iframe里面
            WebDriverWait(self.driver, timeout=timeout).until(ec.frame_to_be_available_and_switch_to_it(frame_content))

    def screen_shot(self, file_name):
        self.driver.get_screenshot_as_file(file_name)

    def cut_ele_img(self,ele,origin_file,reslt_file):
        ele_location = ele.location
        ele_size = ele.size
        stream_obj = Image.open(origin_file)
        stream_obj.crop((ele_location['x'], ele_location['y'], ele_location['x'] + ele_size["width"],
                         ele_location['y'] + ele_size['height'])).save(reslt_file)


    @staticmethod
    def select(webelement, select_type, value):
        """
        选择以何种方式选择下拉框选项
        :param webelement: 下拉框元素
        :param select_type: 选择类型
        :param value: 具体选择的值
        :return:
        """
        if select_type in ("index", "INDEX"):
            Select(webelement).select_by_index(value)
        elif select_type in ("value", "VALUE"):
            Select(webelement).select_by_value(value)
        elif select_type in ("text_value", "TEXT_VALUE"):
            Select(webelement).select_by_visible_text(value)
        else:
            raise NameError("请输入正确的选择类别")

    def open_new_handle(self, url):
        js = 'window.open("%s");' % url
        self.driver.execute_script(js)
        self.driver.switch_to.window(self.driver.window_handles[-1])

    def switch_handle(self, window_name):
        self.driver.switch_to.window(window_name)

    def kill_driver(self):
        self.driver.quit()

    def driver_refresh(self):
        self.driver.refresh()


class OpMail:
    def __init__(self, smtp_user="952618746@qq.com", smtp_password="vmxywawjkgyibbbi"):
        self.smtp_server = smtplib.SMTP("smtp.qq.com", 25)
        self.attachment = MIMEMultipart()
        self.smtp_user = smtp_user
        self.smtp_server.login(user=smtp_user, password=smtp_password)

    def set_base_info(self, sender, receiver, subject, text, text_type="plain", charset="gbk"):
        self.attachment["From"] = sender
        self.attachment["To"] = ",".join(receiver)
        self.attachment["Subject"] = subject
        self.attachment.attach(MIMEText(_text=text, _subtype=text_type, _charset=charset))

    def add_attr(self, filepath):
        attr = MIMEApplication(open(file=filepath, mode="rb").read(), )
        attr.add_header("Content-Disposition", 'attachment', filename=filepath.split("\\")[-1])
        self.attachment.attach(attr,)

    def send_eail(self, receiver):
        self.smtp_server.sendmail(from_addr=self.smtp_user, to_addrs=receiver, msg=self.attachment.as_string())


class OpPicture:

    @staticmethod
    def image_contrast(img1, img2):
        """
        该方法用于比较两张图片的相似度
        :param img1: 图片1的路径
        :param img2: 图片2的路径
        :return:
        """
        image1 = Image.open(img1)
        image2 = Image.open(img2)

        h1 = image1.histogram()
        h2 = image2.histogram()

        result = math.sqrt(reduce(operator.add, list(map(lambda a, b: (a - b) ** 2, h1, h2))) / len(h1))
        return result


class OpUrl:
    def __init__(self,url):
        self.url = url

    def url_decode(self):
        """
        返回解码后的url
        :return:
        """
        return urllib.parse.unquote(self.url)

    def get_url_args(self):
        """
        获取url中的参数
        :return:
        """
        return urllib.parse.parse_qs(self.url)


def my_log():
    """
    生成日志
    :return:
    """
    logger = logging.Logger("YiQiaUiTest")
    os.chdir(os.curdir)
    hander = logging.FileHandler(os.getcwd().split("YiQiaUiTest")[0]+
        r"YiQiaUiTest\data\log\Work_Order_log_%s.log" % time.strftime("%Y_%m_%d", time.localtime()),
        mode="a", encoding="utf8")
    hander.setLevel(level=logging.INFO)
    format_ = logging.Formatter("%(filename)s %(funcName)s %(levelname)s %(lineno)d %(message)s %(asctime)s")
    hander.setFormatter(format_)
    logger.addHandler(hander)
    return logger


class OpExcel:

    def __init__(self,file_name):
        self.file_name = file_name
        self.wb = openpyxl.load_workbook(self.file_name)
        self.ws = self.wb.active

    def get_excel_data(self,sheet_name,rows,columns,json_columns):
        self.ws = self.wb[sheet_name]
        self.ws = list(self.ws.rows)
        need_rows = []
        self.result = []
        for each_row in rows:
            need_rows.append(self.ws[each_row])
        for each_row in need_rows:
            self.result.extend([[each_row[each_column].value for each_column in columns] +
                          [json.loads(each_row[each_column].value) for each_column in json_columns]])
        data = []
        for ecah_test_case in self.result:
            data.append(ecah_test_case[:-1] + list(ecah_test_case[-1].values()))
        return data

if __name__ == '__main__':
    print(OpPicture().image_contrast('D:\YiQiaUiTest\data\Common_test_files\客服端截图比较图片_(jpg_file).png',
                               'D:\YiQiaUiTest\data\screenshot_img\chat_img\目标图片截图_2021_02_19_17_43_15.png'))

    # wb = openpyxl.load_workbook(r'D:\YiQiaUiTest\data\test_case\chat_test_cases.xlsx')
    # ws =wb.active
    # print(list(ws.rows))
    # for each in
